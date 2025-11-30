"""
Internal Audit Analysis Script for Purchase Orders and Contracts
UPDATED VERSION: No Currency Conversion - Use PO amounts directly for foreign currency
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import os
from datetime import datetime, timedelta
import warnings
warnings.filterwarnings('ignore')

# ============================================================================
# CONFIGURATION
# ============================================================================
BASE_PATH = r"D:\DKC_Projects\05_AurionPro\InputFiles"
OUTPUT_FILE = os.path.join(BASE_PATH, f"Audit_Analysis_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

# File names
PO_FILE = "PO Wise Product Report (2).xlsx"
PRS_FILE = "Prs Report (2).xlsx"
CONTRACTS_FILE = "Live Contracts F.Y 2025-2026 Master_Copy.xlsx"

# Company name standardization mapping (keeping your existing mapping)
COMPANY_NAME_MAPPING = {
    'ICICI Bank': 'ICICI Bank Ltd',
    'ICICI Bank ltd': 'ICICI Bank Ltd',
    'State Bank of India': 'State Bank of India',
    'State Bank Of India': 'State Bank of India',
    'State bank of India': 'State Bank of India',
    'State Bank of India - Parlament': 'State Bank of India - Parliament',
    'State Bank of India - Allahabad': 'State Bank of India - Allahabad',
    'State bank of India - CMP': 'State Bank of India - CMP',
    'State Bank of India - Kanpur Metro': 'State Bank of India - Kanpur Metro',
    'State Bank of India - Muree': 'State Bank of India - Muree',
    'State Bank of India ( CDAMS )': 'State Bank of India - CDAMS',
    'Abu Dhabi Commercial Bank': 'Abu Dhabi Commercial Bank',
    'Aurionpro Africa Ltd - Banking': 'Aurionpro Africa Ltd - Banking',
    'Aurionpro Fintech Inc': 'Aurionpro Fintech Inc',
    'Aurionpro Market System Pte Ltd': 'Aurionpro Market System Pte Ltd',
    'Aurionpro Payments Solution Pvt Ltd': 'Aurionpro Payments Solution Pvt Ltd',
    'Aurionpro Solutions Pte Ltd': 'Aurionpro Solutions Pte Ltd',
    'Aurionpro Toshi Automatic Systemv Pvt ltd': 'Aurionpro Toshi Automatic System Pvt Ltd',
    'Aurionpro Transit Pte Ltd': 'Aurionpro Transit Pte Ltd',
    'AXIS Bank Ltd': 'Axis Bank Ltd',
    'Axis Bank Ltd': 'Axis Bank Ltd',
    'Bank of Ceylon': 'Bank of Ceylon',
    'Commercial Bank of Ceylon PLC': 'Commercial Bank of Ceylon PLC',
    'Commercial Bank of Ceylon PLC - Bangakadesh': 'Commercial Bank of Ceylon PLC - Bangladesh',
    'DCB Bank': 'DCB Bank',
    'DFCU Bank': 'DFCU Bank',
    'Punjab National Bank': 'Punjab National Bank',
    'Canara Bank Ltd': 'Canara Bank Ltd',
    'Nations Trust Bank': 'Nations Trust Bank',
    'Union Bank of Colombo PLC': 'Union Bank of Colombo PLC',
    'Yes Bank Ltd': 'Yes Bank Ltd',
    'STC Bank': 'STC Bank',
    'The Saudi Investment Bank': 'The Saudi Investment Bank',
    'Maharashtra Metro Rail Corporation Ltd - 5D BIM': 'Maharashtra Metro Rail Corporation Ltd - 5D BIM',
    'Maharashtra Metro Rail Corporation Ltd - BIM DOT': 'Maharashtra Metro Rail Corporation Ltd - BIM DOT',
    'Maharashtra Metro Rail Corporation Ltd (CIDCO)': 'Maharashtra Metro Rail Corporation Ltd - CIDCO',
    'Delhi Metro Rail Corporation Ltd': 'Delhi Metro Rail Corporation Ltd',
    'Chennai Metro Rail Ltd': 'Chennai Metro Rail Ltd',
    'Nagpur Metro Rail Corporations Ltd': 'Nagpur Metro Rail Corporations Ltd',
    'Noida Rail Corporation Ltd-AFC': 'Noida Rail Corporation Ltd - AFC',
    'Integro Technologies Pte Ltd, Singapore': 'Integro Technologies Pte Ltd - Singapore',
    'Rajcomp Info Services Ltd , 3D City': 'Rajcomp Info Services Ltd - 3D City',
    'SBI Cards and Payment Services Ltd': 'SBI Cards and Payment Services Ltd',
    'WE-3 Supply Chain Services Pvt Ltd': 'WE-3 Supply Chain Services Pvt Ltd',
    'Tourism Department Government of pradesh': 'Tourism Department Government of Pradesh',
    'Uttar Pradesh State Road Tranport Corp': 'Uttar Pradesh State Road Transport Corp',
}

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def get_currency_note(currency):
    """Get a note about currency - NO CONVERSION, just informational"""
    if pd.isna(currency):
        return ""
    
    currency = str(currency).strip()
    
    if currency == 'INR':
        return ""
    
    return f"[Currency: {currency} - Using PO amounts directly]"

def standardize_company_name(name):
    """Standardize company names using mapping"""
    if pd.isna(name):
        return name
    
    name_str = str(name).strip()
    
    # Direct mapping
    if name_str in COMPANY_NAME_MAPPING:
        return COMPANY_NAME_MAPPING[name_str]
    
    # Case-insensitive matching
    for key, value in COMPANY_NAME_MAPPING.items():
        if name_str.lower() == key.lower():
            return value
    
    return name_str

def convert_to_numeric(series, column_name):
    """Convert series to numeric, handling errors"""
    try:
        if series.dtype == 'object':
            series = series.astype(str).str.replace(',', '').str.replace('₹', '').str.strip()
        
        numeric_series = pd.to_numeric(series, errors='coerce')
        numeric_series = numeric_series.fillna(0)
        
        return numeric_series
    except Exception as e:
        print(f"  Warning: Could not convert {column_name} to numeric: {str(e)}")
        return series.fillna(0)

def standardize_status(status):
    """Standardize status values to handle case variations"""
    if pd.isna(status):
        return 'Unknown'
    
    status_str = str(status).strip().upper()
    
    # Map variations to standard values
    if status_str in ['APPROVED', 'APPROVE']:
        return 'Approved'
    elif status_str in ['PENDING', 'PEND']:
        return 'Pending'
    elif status_str in ['REJECTED', 'REJECT']:
        return 'Rejected'
    else:
        return status_str.title()

def apply_header_style(ws, header_row=1):
    """Apply consistent header styling"""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def auto_adjust_column_width(ws):
    """Auto-adjust column widths"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def add_borders(ws, start_row=1, end_row=None, start_col=1, end_col=None):
    """Add borders to cells"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if end_row is None:
        end_row = ws.max_row
    if end_col is None:
        end_col = ws.max_column
        
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = thin_border

# ============================================================================
# DATA LOADING FUNCTIONS
# ============================================================================

def load_po_data():
    """Load Purchase Order data with proper data type handling"""
    print("Loading PO Wise Product Report...")
    po_path = os.path.join(BASE_PATH, PO_FILE)
    df = pd.read_excel(po_path, sheet_name='RawData')
    
    # Convert numeric columns
    numeric_columns = ['Rate', 'Quatity', 'Amount', 'CGST', 'SGST', 'IGST', 'Total Discount', 'Final Amt.']
    for col in numeric_columns:
        if col in df.columns:
            df[col] = convert_to_numeric(df[col], col)
    
    # Standardize vendor names
    if 'Vendor' in df.columns:
        df['Vendor'] = df['Vendor'].apply(standardize_company_name)
    
    # PO data is assumed to be in INR
    df['Currency'] = 'INR'
    df['Final Amt. (INR)'] = df['Final Amt.']
    
    print(f"  Loaded {len(df)} PO line items")
    print(f"  Unique POs: {df['PO. No.'].nunique()}")
    return df

def load_prs_data():
    """Load PRS data - NO CURRENCY CONVERSION, use amounts as-is"""
    print("Loading PRS Report...")
    prs_path = os.path.join(BASE_PATH, PRS_FILE)
    df = pd.read_excel(prs_path, sheet_name='RawData')
    
    # Convert numeric columns
    numeric_columns = ['Invoice Amount', 'PO Amt w/o Tax', 'PO Tax', 'PO Total']
    for col in numeric_columns:
        if col in df.columns:
            df[col] = convert_to_numeric(df[col], col)
    
    # Standardize vendor names
    if 'Vendor' in df.columns:
        df['Vendor'] = df['Vendor'].apply(standardize_company_name)
    
    # Standardize status
    if 'Status' in df.columns:
        df['Status'] = df['Status'].apply(standardize_status)
    
    # Handle Currency column
    if 'Currency' not in df.columns:
        df['Currency'] = 'INR'
        print("  Warning: No Currency column found, assuming INR for all transactions")
    else:
        df['Currency'] = df['Currency'].fillna('INR')
    
    # Store original values - NO CONVERSION TO INR
    df['Invoice Amount (Original)'] = df['Invoice Amount']
    df['PO Total AsPer PRS_Record'] = df['PO Total']  # CHANGED: Renamed column
    df['Original Currency'] = df['Currency']
    
    # For INR transactions, use amounts as-is
    # For non-INR transactions, we'll use PO amounts directly in analysis
    
    # Add currency notes
    df['Currency Note'] = df['Currency'].apply(get_currency_note)
    
    # Convert date columns
    if 'Payment Due Date' in df.columns:
        df['Payment Due Date'] = pd.to_datetime(df['Payment Due Date'], errors='coerce')
    
    # Show currency breakdown
    currency_summary = df.groupby('Currency').agg({
        'PRS No': 'count',
        'Invoice Amount (Original)': 'sum',
        'PO Total AsPer PRS_Record': 'sum'
    }).reset_index()
    
    print("\n  Currency Breakdown (Amounts in Original Currency):")
    for _, row in currency_summary.iterrows():
        print(f"    {row['Currency']}: {row['PRS No']} transactions, " +
              f"Invoice Amount: {row['Invoice Amount (Original)']:,.2f}, " +
              f"PO Total: {row['PO Total AsPer PRS_Record']:,.2f}")
    
    print(f"\n  Loaded {len(df)} PRS records (payment installments)")
    print(f"  Unique POs in PRS: {df['PO No.'].nunique()}")
    return df

def load_contracts_data():
    """Load Live Contracts data from all sheets with standardization"""
    print("Loading Live Contracts data...")
    contracts_path = os.path.join(BASE_PATH, CONTRACTS_FILE)
    
    sheets_to_load = ["Banking 25-26", "GSG 25-26", "GSG 25-26 PSS", "Others 25-26"]
    contracts_dict = {}
    
    for sheet in sheets_to_load:
        try:
            df = pd.read_excel(contracts_path, sheet_name=sheet)
            
            # Standardize client names
            if 'Name of the Clients' in df.columns:
                df['Name of the Clients'] = df['Name of the Clients'].apply(standardize_company_name)
                df['Client Name (Standardized)'] = df['Name of the Clients']
            
            # Convert numeric columns
            for col in df.columns:
                if any(keyword in str(col).lower() for keyword in ['amt', 'amount', 'po', 'invoice', 'balance', 'total']):
                    if df[col].dtype == 'object':
                        df[col] = convert_to_numeric(df[col], col)
            
            contracts_dict[sheet] = df
            print(f"  Loaded {len(df)} records from {sheet}")
        except Exception as e:
            print(f"  Warning: Could not load {sheet}: {str(e)}")
    
    return contracts_dict

# ============================================================================
# ANALYSIS FUNCTIONS
# ============================================================================

def analysis_1_po_vs_prs_matching(po_df, prs_df, wb):
    """
    Analysis 1: PO vs PRS Matching - SMART CURRENCY-AWARE COMPARISON
    
    Changes:
    1. Detects multi-currency PRS entries for same PO
    2. Prioritizes INR values from PRS for comparison
    3. Flags mixed-currency POs separately
    4. Provides detailed currency breakdown per PO
    """
    print("\nPerforming Analysis 1: PO vs PRS Matching (SMART CURRENCY HANDLING)...")
    
    # Aggregate PO data (always in INR)
    po_summary = po_df.groupby('PO. No.').agg({
        'Project': 'first',
        'Project Pin': 'first',
        'Vendor': 'first',
        'Order Dt.': 'first',
        'Final Amt. (INR)': 'sum'
    }).reset_index()
    
    # NEW: Analyze PRS currency structure per PO
    prs_currency_analysis = prs_df.groupby('PO No.')['Currency'].apply(
        lambda x: {
            'currencies': list(x.unique()),
            'has_inr': 'INR' in x.values,
            'is_multi': len(x.unique()) > 1,
            'currency_list': ', '.join(sorted(x.unique()))
        }
    ).to_dict()
    
    # NEW: Smart aggregation of PRS data with currency awareness
    prs_summary_list = []
    
    for po_no, po_group in prs_df.groupby('PO No.'):
        currency_info = prs_currency_analysis.get(po_no, {})
        
        # Determine which PO Total to use for comparison
        if currency_info.get('has_inr', False):
            # If INR entries exist, use INR PO Total for comparison
            inr_entries = po_group[po_group['Currency'] == 'INR']
            po_total_for_comparison = inr_entries['PO Total'].iloc[0] if len(inr_entries) > 0 else 0
            comparison_currency = 'INR'
        else:
            # No INR entries - use first entry but mark for manual review
            po_total_for_comparison = po_group['PO Total'].iloc[0]
            comparison_currency = po_group['Currency'].iloc[0]
        
        # Calculate total invoiced - separate by currency
        invoice_by_currency = po_group.groupby('Currency')['Invoice Amount'].sum().to_dict()
        total_invoiced_inr = invoice_by_currency.get('INR', 0)
        
        # Create detailed currency breakdown
        currency_breakdown = []
        for curr in sorted(po_group['Currency'].unique()):
            curr_data = po_group[po_group['Currency'] == curr]
            curr_invoiced = curr_data['Invoice Amount'].sum()
            curr_po_total = curr_data['PO Total'].iloc[0]
            currency_breakdown.append(f"{curr}: PO={curr_po_total:,.2f}, Inv={curr_invoiced:,.2f}")
        
        prs_summary_list.append({
            'PO No.': po_no,
            'PO Total (For Comparison)': po_total_for_comparison,
            'Comparison Currency': comparison_currency,
            'Total Invoiced (INR)': total_invoiced_inr,
            'Total Invoiced (All Currencies)': po_group['Invoice Amount'].sum(),
            'Status': ', '.join(po_group['Status'].dropna().astype(str).unique()),
            'No. of Installments': len(po_group),
            'Currencies Used': currency_info.get('currency_list', 'Unknown'),
            'Is Multi-Currency': currency_info.get('is_multi', False),
            'Has INR Entry': currency_info.get('has_inr', False),
            'Currency Breakdown': ' | '.join(currency_breakdown)
        })
    
    prs_summary = pd.DataFrame(prs_summary_list)
    
    # Merge with PO data
    analysis = po_summary.merge(
        prs_summary, 
        left_on='PO. No.', 
        right_on='PO No.',
        how='left'
    )
    
    # Fill missing values
    analysis['PO Total (For Comparison)'] = analysis['PO Total (For Comparison)'].fillna(0)
    analysis['Comparison Currency'] = analysis['Comparison Currency'].fillna('INR')
    analysis['Total Invoiced (INR)'] = analysis['Total Invoiced (INR)'].fillna(0)
    analysis['Total Invoiced (All Currencies)'] = analysis['Total Invoiced (All Currencies)'].fillna(0)
    analysis['No. of Installments'] = analysis['No. of Installments'].fillna(0).astype(int)
    analysis['Status'] = analysis['Status'].fillna('No PRS Record')
    analysis['Currencies Used'] = analysis['Currencies Used'].fillna('INR')
    analysis['Is Multi-Currency'] = analysis['Is Multi-Currency'].fillna(False)
    analysis['Has INR Entry'] = analysis['Has INR Entry'].fillna(False)
    analysis['Currency Breakdown'] = analysis['Currency Breakdown'].fillna('')
    
    # NEW: Smart calculation logic based on currency situation
    def calculate_balance_and_fulfillment(row):
        """Calculate balance and fulfillment based on currency situation"""
        
        # Case 1: Pure INR PO with INR PRS entries
        if row['Comparison Currency'] == 'INR' and not row['Is Multi-Currency']:
            balance = row['Final Amt. (INR)'] - row['Total Invoiced (INR)']
            if row['Final Amt. (INR)'] > 0:
                fulfillment = round((row['Total Invoiced (INR)'] / row['Final Amt. (INR)']) * 100, 2)
            else:
                fulfillment = 0
            matching_status = 'Pure INR - OK'
            
        # Case 2: Multi-currency PO but has INR entry for comparison
        elif row['Comparison Currency'] == 'INR' and row['Is Multi-Currency']:
            balance = row['Final Amt. (INR)'] - row['Total Invoiced (INR)']
            if row['Final Amt. (INR)'] > 0:
                fulfillment = round((row['Total Invoiced (INR)'] / row['Final Amt. (INR)']) * 100, 2)
            else:
                fulfillment = 0
            matching_status = 'Multi-Currency with INR'
            
        # Case 3: Pure foreign currency - no INR comparison possible
        elif row['Comparison Currency'] != 'INR' and not row['Is Multi-Currency']:
            balance = None
            fulfillment = None
            matching_status = f'Foreign Currency Only ({row["Comparison Currency"]})'
            
        # Case 4: Multi-currency without INR entry
        else:
            balance = None
            fulfillment = None
            matching_status = 'Multi-Currency - Manual Review Required'
        
        return pd.Series({
            'Balance Pending (INR)': balance,
            'Fulfillment %': fulfillment,
            'Currency Matching Status': matching_status
        })
    
    # Apply calculation
    calc_results = analysis.apply(calculate_balance_and_fulfillment, axis=1)
    analysis = pd.concat([analysis, calc_results], axis=1)
    
    # NEW: Enhanced PO matching check - only for comparable situations
    def check_po_matching(row):
        """Check PO total matching with smart currency handling"""
        
        if row['Status'] == 'No PRS Record':
            return 'No PRS Record Found'
        
        # Only compare when currencies match
        if row['Comparison Currency'] == 'INR':
            diff = abs(row['Final Amt. (INR)'] - row['PO Total (For Comparison)'])
            if diff > 1000:
                return f'Amount Mismatch: ₹{diff:,.2f} difference'
            else:
                return 'Amounts Match'
        else:
            return f'Cannot Compare - PRS in {row["Comparison Currency"]}'
    
    analysis['PO Matching Check'] = analysis.apply(check_po_matching, axis=1)
    
    # Add detailed notes for auditors
    def create_audit_note(row):
        """Create detailed audit notes"""
        notes = []
        
        if row['Is Multi-Currency']:
            notes.append(f"⚠️ MULTI-CURRENCY PO: {row['Currencies Used']}")
        
        if not row['Has INR Entry'] and row['Comparison Currency'] != 'INR':
            notes.append(f"🔍 NO INR ENTRY - All entries in {row['Comparison Currency']}")
        
        if row['Is Multi-Currency'] and row['Has INR Entry']:
            notes.append("✓ Mixed currencies but INR entry found for comparison")
        
        if row['Currency Breakdown']:
            notes.append(f"Details: {row['Currency Breakdown']}")
        
        return ' | '.join(notes) if notes else ''
    
    analysis['Audit Notes'] = analysis.apply(create_audit_note, axis=1)
    
    # Reorder columns for clarity
    analysis = analysis[[
        'PO. No.', 'Project', 'Project Pin', 'Vendor', 'Order Dt.',
        'Final Amt. (INR)', 
        'PO Total (For Comparison)', 'Comparison Currency',
        'Total Invoiced (INR)', 'Total Invoiced (All Currencies)',
        'Balance Pending (INR)', 'Fulfillment %', 
        'No. of Installments', 'Status',
        'Currencies Used', 'Is Multi-Currency', 'Has INR Entry',
        'Currency Matching Status', 'PO Matching Check',
        'Currency Breakdown', 'Audit Notes'
    ]]
    
    # Sort by Final Amount
    analysis = analysis.sort_values('Final Amt. (INR)', ascending=False)
    
    # Create worksheet
    ws = wb.create_sheet("1_PO_vs_PRS_Matching")
    
    # Write data
    for r in dataframe_to_rows(analysis, index=False, header=True):
        ws.append(r)
    
    # Add comprehensive summary
    data_end_row = len(analysis) + 1
    formula_row = data_end_row + 2
    
    ws[f'A{formula_row}'] = "FINANCIAL SUMMARY (Smart Currency Handling)"
    ws[f'A{formula_row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{formula_row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells(f'A{formula_row}:E{formula_row}')
    
    formula_row += 2
    ws[f'A{formula_row}'] = "Total PO Amount (INR):"
    ws[f'F{formula_row}'] = f'=SUM(F2:F{data_end_row})'
    ws[f'F{formula_row}'].number_format = '#,##0.00'
    
    ws[f'A{formula_row+1}'] = "Total Invoiced (INR only):"
    ws[f'I{formula_row+1}'] = f'=SUM(I2:I{data_end_row})'
    ws[f'I{formula_row+1}'].number_format = '#,##0.00'
    
    ws[f'A{formula_row+2}'] = "Total Invoiced (All Currencies):"
    ws[f'J{formula_row+2}'] = f'=SUM(J2:J{data_end_row})'
    ws[f'J{formula_row+2}'].number_format = '#,##0.00'
    
    ws[f'A{formula_row+3}'] = "Total Balance Pending (INR only):"
    ws[f'K{formula_row+3}'] = f'=SUMIF(H2:H{data_end_row},"INR",K2:K{data_end_row})'
    ws[f'K{formula_row+3}'].number_format = '#,##0.00'
    
    ws[f'A{formula_row+4}'] = "Overall Fulfillment % (INR only):"
    ws[f'L{formula_row+4}'] = f'=IF(F{formula_row}=0,0,I{formula_row+1}/F{formula_row}*100)'
    ws[f'L{formula_row+4}'].number_format = '0.00"%"'
    
    # Currency breakdown summary
    formula_row += 6
    ws[f'A{formula_row}'] = "PO CURRENCY BREAKDOWN"
    ws[f'A{formula_row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{formula_row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.merge_cells(f'A{formula_row}:D{formula_row}')
    
    formula_row += 1
    ws[f'A{formula_row}'] = "Pure INR POs:"
    ws[f'C{formula_row}'] = f'=COUNTIF(R2:R{data_end_row},"Pure INR - OK")'
    
    formula_row += 1
    ws[f'A{formula_row}'] = "Multi-Currency with INR:"
    ws[f'C{formula_row}'] = f'=COUNTIF(R2:R{data_end_row},"Multi-Currency with INR")'
    
    formula_row += 1
    ws[f'A{formula_row}'] = "Foreign Currency Only:"
    ws[f'C{formula_row}'] = f'=COUNTIF(P2:P{data_end_row},TRUE)-COUNTIF(Q2:Q{data_end_row},TRUE)+COUNTIF(R2:R{data_end_row},"Foreign Currency Only*")'
    
    formula_row += 1
    ws[f'A{formula_row}'] = "Multi-Currency (No INR):"
    ws[f'C{formula_row}'] = f'=COUNTIF(R2:R{data_end_row},"Multi-Currency - Manual Review Required")'
    
    formula_row += 1
    ws[f'A{formula_row}'] = "No PRS Record:"
    ws[f'C{formula_row}'] = f'=COUNTIF(N2:N{data_end_row},"No PRS Record")'
    
    # Matching status summary
    formula_row += 2
    ws[f'A{formula_row}'] = "PO MATCHING STATUS"
    ws[f'A{formula_row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{formula_row}'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws.merge_cells(f'A{formula_row}:D{formula_row}')
    
    formula_row += 1
    ws[f'A{formula_row}'] = "Amounts Match:"
    ws[f'C{formula_row}'] = f'=COUNTIF(S2:S{data_end_row},"Amounts Match")'
    
    formula_row += 1
    ws[f'A{formula_row}'] = "Amount Mismatches:"
    ws[f'C{formula_row}'] = f'=COUNTIF(S2:S{data_end_row},"Amount Mismatch*")'
    
    formula_row += 1
    ws[f'A{formula_row}'] = "Cannot Compare (Foreign Currency):"
    ws[f'C{formula_row}'] = f'=COUNTIF(S2:S{data_end_row},"Cannot Compare*")'
    
    apply_header_style(ws)
    auto_adjust_column_width(ws)
    add_borders(ws)
    
    # Enhanced highlighting
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    orange_fill = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
    
    for row in range(2, data_end_row + 1):
        currency_status = ws[f'R{row}'].value
        matching_check = ws[f'S{row}'].value
        is_multi = ws[f'P{row}'].value
        
        try:
            # Highlight based on currency situation
            if currency_status and 'Pure INR' in str(currency_status):
                # Check fulfillment for pure INR
                fulfillment = ws[f'L{row}'].value
                if fulfillment and fulfillment < 50:
                    ws[f'L{row}'].fill = red_fill
                elif fulfillment and fulfillment >= 100:
                    ws[f'L{row}'].fill = green_fill
                    
            elif currency_status and 'Multi-Currency with INR' in str(currency_status):
                # Highlight multi-currency rows in light orange
                for col in ['O', 'P', 'Q', 'R', 'T', 'U']:
                    ws[f'{col}{row}'].fill = orange_fill
                    
            elif currency_status and 'Foreign Currency Only' in str(currency_status):
                # Highlight pure foreign currency in blue
                for col in ['G', 'H', 'I', 'J', 'K', 'L', 'O', 'R', 'S', 'U']:
                    ws[f'{col}{row}'].fill = blue_fill
                    
            elif currency_status and 'Manual Review Required' in str(currency_status):
                # Highlight manual review needed in yellow
                for col in ['O', 'P', 'Q', 'R', 'S', 'T', 'U']:
                    ws[f'{col}{row}'].fill = yellow_fill
            
            # Highlight matching issues
            if matching_check and 'Mismatch' in str(matching_check):
                ws[f'S{row}'].fill = red_fill
                ws[f'G{row}'].fill = red_fill
                
        except:
            pass
    
    print(f"  ✓ Analyzed {len(analysis)} POs with SMART currency handling")
    
    # Print currency statistics
    pure_inr = len(analysis[analysis['Currency Matching Status'] == 'Pure INR - OK'])
    multi_with_inr = len(analysis[analysis['Currency Matching Status'] == 'Multi-Currency with INR'])
    foreign_only = len(analysis[analysis['Currency Matching Status'].str.contains('Foreign Currency Only', na=False)])
    manual_review = len(analysis[analysis['Currency Matching Status'] == 'Multi-Currency - Manual Review Required'])
    
    print(f"  ✓ Pure INR POs: {pure_inr}")
    print(f"  ✓ Multi-Currency with INR: {multi_with_inr}")
    print(f"  ✓ Foreign Currency Only: {foreign_only}")
    print(f"  ⚠️ Manual Review Required: {manual_review}")

def analysis_2_vendor_wise_summary(po_df, prs_df, wb):
    """Analysis 2: Vendor-wise Summary - INR only"""
    print("\nPerforming Analysis 2: Vendor-wise Summary (INR only)...")
    
    # PO summary
    po_vendor = po_df.groupby('Vendor').agg({
        'PO. No.': 'nunique',
        'Final Amt. (INR)': 'sum',
        'Product': lambda x: x.nunique()
    }).reset_index()
    po_vendor.columns = ['Vendor', 'No. of POs', 'Total PO Amount (INR)', 'Unique Products']
    
    # PRS summary - INR transactions only
    prs_inr = prs_df[prs_df['Currency'] == 'INR'].copy()
    prs_grouped = prs_inr.groupby(['Vendor', 'PO No.']).agg({
        'PO Total AsPer PRS_Record': 'first',
        'Invoice Amount (Original)': 'sum',
        'Status': lambda x: (x == 'Approved').sum()
    }).reset_index()
    
    prs_vendor = prs_grouped.groupby('Vendor').agg({
        'PO Total AsPer PRS_Record': 'sum',
        'Invoice Amount (Original)': 'sum',
        'Status': 'sum'
    }).reset_index()
    prs_vendor.columns = ['Vendor', 'Total PRS Order Value (INR)', 'Total Invoiced (INR)', 'Approved Payments']
    
    # Merge
    vendor_analysis = po_vendor.merge(prs_vendor, on='Vendor', how='left').fillna(0)
    vendor_analysis['Payment Pending (INR)'] = vendor_analysis['Total PO Amount (INR)'] - vendor_analysis['Total Invoiced (INR)']
    vendor_analysis['Payment %'] = np.where(
        vendor_analysis['Total PO Amount (INR)'] > 0,
        (vendor_analysis['Total Invoiced (INR)'] / vendor_analysis['Total PO Amount (INR)'] * 100).round(2),
        0
    )
    
    vendor_analysis = vendor_analysis.sort_values('Total PO Amount (INR)', ascending=False)
    
    # Create worksheet
    ws = wb.create_sheet("2_Vendor_Wise_Summary")
    
    for r in dataframe_to_rows(vendor_analysis, index=False, header=True):
        ws.append(r)
    
    # Add summary
    data_end_row = len(vendor_analysis) + 1
    summary_row = data_end_row + 2
    
    ws[f'A{summary_row}'] = "GRAND TOTAL (INR only)"
    ws[f'A{summary_row}'].font = Font(bold=True, size=12)
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}{summary_row}'] = f'=SUM({col}2:{col}{data_end_row})'
        if col in ['C', 'E', 'F', 'G']:
            ws[f'{col}{summary_row}'].number_format = '#,##0.00'
    
    apply_header_style(ws)
    auto_adjust_column_width(ws)
    add_borders(ws)
    
    print(f"  ✓ Analyzed {len(vendor_analysis)} vendors (INR transactions only)")

def analysis_3_project_wise_summary(po_df, prs_df, wb):
    """Analysis 3: Project-wise Summary - INR only"""
    print("\nPerforming Analysis 3: Project-wise Summary (INR only)...")
    
    project_analysis = po_df.groupby(['Project Pin', 'Project']).agg({
        'PO. No.': 'nunique',
        'Vendor': 'nunique',
        'Final Amt. (INR)': 'sum',
        'Product': 'count'
    }).reset_index()
    
    project_analysis.columns = [
        'Project Pin', 'Project Name', 'No. of POs', 
        'No. of Vendors', 'Total Purchase Amount (INR)', 'Total Items'
    ]
    
    # Get PRS data - INR only
    prs_inr = prs_df[prs_df['Currency'] == 'INR'].copy()
    prs_inr['Project Pin'] = prs_inr['Pin/Project'].str.split(' - ').str[0].str.strip()
    
    prs_grouped = prs_inr.groupby(['Project Pin', 'PO No.']).agg({
        'PO Total AsPer PRS_Record': 'first',
        'Invoice Amount (Original)': 'sum'
    }).reset_index()
    
    prs_project = prs_grouped.groupby('Project Pin').agg({
        'PO Total AsPer PRS_Record': 'sum',
        'Invoice Amount (Original)': 'sum'
    }).reset_index()
    
    project_analysis = project_analysis.merge(
        prs_project, on='Project Pin', how='left'
    ).fillna(0)
    
    project_analysis['Balance Pending (INR)'] = project_analysis['Total Purchase Amount (INR)'] - project_analysis['Invoice Amount (Original)']
    project_analysis['Completion %'] = np.where(
        project_analysis['Total Purchase Amount (INR)'] > 0,
        (project_analysis['Invoice Amount (Original)'] / project_analysis['Total Purchase Amount (INR)'] * 100).round(2),
        0
    )
    
    project_analysis = project_analysis.sort_values('Total Purchase Amount (INR)', ascending=False)
    
    ws = wb.create_sheet("3_Project_Wise_Summary")
    
    for r in dataframe_to_rows(project_analysis, index=False, header=True):
        ws.append(r)
    
    # Add totals
    data_end_row = len(project_analysis) + 1
    summary_row = data_end_row + 2
    
    ws[f'A{summary_row}'] = "TOTAL (INR only)"
    ws[f'A{summary_row}'].font = Font(bold=True)
    for col in ['C', 'E', 'F', 'G', 'H', 'I']:
        ws[f'{col}{summary_row}'] = f'=SUM({col}2:{col}{data_end_row})'
        if col in ['E', 'G', 'H', 'I']:
            ws[f'{col}{summary_row}'].number_format = '#,##0.00'
    
    apply_header_style(ws)
    auto_adjust_column_width(ws)
    add_borders(ws)
    
    print(f"  ✓ Analyzed {len(project_analysis)} projects (INR transactions only)")

def analysis_4_payment_status_tracking(prs_df, wb):
    """Analysis 4: Payment Status Tracking - Original amounts only"""
    print("\nPerforming Analysis 4: Payment Status Tracking (Original amounts only)...")
    
    prs_df['Payment Due Date'] = pd.to_datetime(prs_df['Payment Due Date'], errors='coerce')
    today = pd.Timestamp.now()
    
    # Better audit logic: Different checks based on status
    prs_df['Days Since Due'] = (today - prs_df['Payment Due Date']).dt.days
    prs_df['Days Since Due'] = prs_df['Days Since Due'].fillna(0)
    
    # Create audit flags based on status
    def get_audit_flag(row):
        status = row['Status']
        days = row['Days Since Due']
        
        if status == 'Approved':
            # Approved payments - check if they were approved on time
            if days > 90:
                return 'Long Approval Cycle (90+ days)'
            elif days > 60:
                return 'Delayed Approval (60-90 days)'
            else:
                return 'Normal'
        
        elif status == 'Pending':
            # Pending payments - these are actual concerns
            if days > 30:
                return 'CRITICAL: Pending >30 days'
            elif days > 0:
                return 'Action Required: Overdue'
            else:
                return 'Pending - Not Yet Due'
        
        elif status == 'Rejected':
            return 'Rejected - Requires Review'
        
        else:
            return 'Unknown Status'
    
    prs_df['Audit Flag'] = prs_df.apply(get_audit_flag, axis=1)
    
    # Create analysis dataframe
    payment_analysis = prs_df[[
        'PRS No', 'Pin/Project', 'PO No.', 'Vendor', 'Invoice No.',
        'Invoice Amount (Original)', 'Original Currency', 
        'Payment Due Date', 'Days Since Due', 'Status', 'Audit Flag'
    ]].copy()
    
    # After creating payment_analysis dataframe, add:
    payment_analysis['Amount Display'] = payment_analysis.apply(
        lambda row: f"{row['Original Currency']} {row['Invoice Amount (Original)']:,.2f}" 
        if pd.notna(row['Original Currency']) else f"₹{row['Invoice Amount (Original)']:,.2f}",
        axis=1
    )
    
    # Add aging categories
    def categorize_aging(row):
        days = row['Days Since Due']
        status = row['Status']
        
        if status == 'Approved':
            return 'Approved (Completed)'
        elif days <= 0:
            return 'Not Yet Due'
        elif days <= 30:
            return '1-30 Days Past Due'
        elif days <= 60:
            return '31-60 Days Past Due'
        elif days <= 90:
            return '61-90 Days Past Due'
        else:
            return '90+ Days Past Due'
    
    payment_analysis['Aging Category'] = payment_analysis.apply(categorize_aging, axis=1)
    
    # Create worksheet
    ws = wb.create_sheet("4_Payment_Status_Analysis")
    
    for r in dataframe_to_rows(payment_analysis, index=False, header=True):
        ws.append(r)
    
    # Add comprehensive summary
    data_end_row = len(payment_analysis) + 1
    summary_start_row = data_end_row + 3
    
    ws[f'A{summary_start_row}'] = "PAYMENT STATUS SUMMARY (All currencies)"
    ws[f'A{summary_start_row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{summary_start_row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells(f'A{summary_start_row}:D{summary_start_row}')
    
    summary_start_row += 2
    ws[f'A{summary_start_row}'] = "Status"
    ws[f'B{summary_start_row}'] = "Count"
    ws[f'C{summary_start_row}'] = "Amount (Original)"
    ws[f'D{summary_start_row}'] = "% of Total"
    apply_header_style(ws, summary_start_row)
    
    # Status breakdown
    statuses = ['Approved', 'Pending', 'Rejected']
    for idx, status in enumerate(statuses):
        row = summary_start_row + 1 + idx
        ws[f'A{row}'] = status
        ws[f'B{row}'] = f'=COUNTIF(K2:K{data_end_row},"{status}")'
        ws[f'C{row}'] = f'=SUMIF(K2:K{data_end_row},"{status}",F2:F{data_end_row})'
        ws[f'C{row}'].number_format = '#,##0.00'
        ws[f'D{row}'] = f'=C{row}/SUM(C{summary_start_row+1}:C{summary_start_row+3})*100'
        ws[f'D{row}'].number_format = '0.00"%"'
    
    # Audit flags summary
    audit_row = summary_start_row + 6
    ws[f'A{audit_row}'] = "AUDIT PRIORITY FLAGS"
    ws[f'A{audit_row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{audit_row}'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws.merge_cells(f'A{audit_row}:D{audit_row}')
    
    audit_row += 2
    ws[f'A{audit_row}'] = "Audit Flag"
    ws[f'B{audit_row}'] = "Count"
    ws[f'C{audit_row}'] = "Amount (Original)"
    apply_header_style(ws, audit_row)
    
    # Get unique audit flags
    critical_flags = [
        'CRITICAL: Pending >30 days',
        'Action Required: Overdue',
        'Rejected - Requires Review',
        'Long Approval Cycle (90+ days)'
    ]
    
    for idx, flag in enumerate(critical_flags):
        row = audit_row + 1 + idx
        ws[f'A{row}'] = flag
        ws[f'B{row}'] = f'=COUNTIF(L2:L{data_end_row},"{flag}")'
        ws[f'C{row}'] = f'=SUMIF(L2:L{data_end_row},"{flag}",F2:F{data_end_row})'
        ws[f'C{row}'].number_format = '#,##0.00'
    
    apply_header_style(ws)
    auto_adjust_column_width(ws)
    add_borders(ws)
    
    # Highlight by audit flag
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    for row in range(2, data_end_row + 1):
        flag = ws[f'L{row}'].value
        if flag and 'CRITICAL' in str(flag):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = red_fill
        elif flag and 'Action Required' in str(flag):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = orange_fill
        elif flag and 'Normal' in str(flag):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = green_fill
    
    print(f"  ✓ Analyzed {len(payment_analysis)} payment records with original amounts")

def analysis_5_tax_analysis(po_df, prs_df, wb):
    """Analysis 5: Tax Analysis - INR only"""
    print("\nPerforming Analysis 5: Tax Analysis (INR only)...")
    
    tax_analysis = po_df.groupby(['PO. No.', 'Vendor', 'Project']).agg({
        'Amount': 'sum',
        'CGST': 'sum',
        'SGST': 'sum',
        'IGST': 'sum',
        'Total Discount': 'sum',
        'Final Amt. (INR)': 'sum'
    }).reset_index()
    
    tax_analysis['Total Tax (Calculated)'] = (
        tax_analysis['CGST'] + 
        tax_analysis['SGST'] + 
        tax_analysis['IGST']
    )
    
    # PRS tax data - INR only
    prs_inr = prs_df[prs_df['Currency'] == 'INR'].copy()
    prs_tax = prs_inr.groupby('PO No.').agg({
        'PO Tax': 'first'
    }).reset_index()
    
    tax_analysis = tax_analysis.merge(
        prs_tax,
        left_on='PO. No.',
        right_on='PO No.',
        how='left'
    ).fillna(0)
    
    # Convert PO Tax to numeric
    tax_analysis['PO Tax'] = convert_to_numeric(tax_analysis['PO Tax'], 'PO Tax')
    
    tax_analysis['Tax Difference'] = tax_analysis['Total Tax (Calculated)'] - tax_analysis['PO Tax']
    tax_analysis['Tax Verification'] = np.where(
        abs(tax_analysis['Tax Difference']) > 100,
        'Review Required',
        'OK'
    )
    
    ws = wb.create_sheet("5_Tax_Analysis")
    
    for r in dataframe_to_rows(tax_analysis, index=False, header=True):
        ws.append(r)
    
    # Add summary
    data_end_row = len(tax_analysis) + 1
    summary_row = data_end_row + 2
    
    ws[f'A{summary_row}'] = "TOTAL (INR only)"
    ws[f'A{summary_row}'].font = Font(bold=True)
    for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        ws[f'{col}{summary_row}'] = f'=SUM({col}2:{col}{data_end_row})'
        ws[f'{col}{summary_row}'].number_format = '#,##0.00'
    
    apply_header_style(ws)
    auto_adjust_column_width(ws)
    add_borders(ws)
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for row in range(2, data_end_row + 1):
        if ws[f'M{row}'].value == 'Review Required':
            ws[f'M{row}'].fill = yellow_fill
            ws[f'L{row}'].fill = yellow_fill
    
    print(f"  ✓ Analyzed tax for {len(tax_analysis)} POs (INR only)")

def analysis_6_audit_exceptions(po_df, prs_df, wb):
    """Analysis 6: Audit Exceptions - SMART CURRENCY-AWARE DETECTION"""
    print("\nPerforming Analysis 6: Audit Exceptions (SMART CURRENCY HANDLING)...")
    
    exceptions = []
    
    # Exception 1: POs without PRS records
    po_nos = set(po_df['PO. No.'].unique())
    prs_nos = set(prs_df['PO No.'].unique())
    missing_prs = po_nos - prs_nos
    
    for po_no in missing_prs:
        po_match = po_df[po_df['PO. No.'] == po_no]
        if len(po_match) > 0:
            po_record = po_match.iloc[0]
            exceptions.append({
                'Exception Type': 'PO without Payment Tracking',
                'Severity': 'High',
                'PO No.': po_no,
                'Project': po_record['Project'],
                'Vendor': po_record['Vendor'],
                'Amount (INR)': po_record['Final Amt. (INR)'],
                'Currency Info': 'INR',
                'Details': 'No payment tracking record found - Order placed but no payment records'
            })
    
    # Exception 2: PO Total mismatch - SMART CURRENCY HANDLING
    po_summary = po_df.groupby('PO. No.').agg({
        'Final Amt. (INR)': 'sum',
        'Project': 'first',
        'Vendor': 'first'
    }).reset_index()
    
    # NEW: Get currency-aware PO Total for comparison
    prs_currency_analysis = prs_df.groupby('PO No.').apply(
        lambda x: pd.Series({
            'currencies': ', '.join(sorted(x['Currency'].unique())),
            'has_inr': 'INR' in x['Currency'].values,
            'is_multi': len(x['Currency'].unique()) > 1,
            'po_total_inr': x[x['Currency'] == 'INR']['PO Total'].iloc[0] if 'INR' in x['Currency'].values else None,
            'po_total_first': x['PO Total'].iloc[0],
            'currency_first': x['Currency'].iloc[0]
        })
    ).reset_index()
    
    merged = po_summary.merge(prs_currency_analysis, left_on='PO. No.', right_on='PO No.', how='inner')
    
    # Only check mismatches for comparable situations (INR to INR)
    comparable = merged[merged['has_inr'] == True].copy()
    comparable['Difference'] = abs(comparable['Final Amt. (INR)'] - comparable['po_total_inr'])
    
    # Higher threshold for mismatches (1000 INR)
    significant_diff = comparable[comparable['Difference'] > 1000]
    
    for _, row in significant_diff.iterrows():
        exceptions.append({
            'Exception Type': 'PO Amount Mismatch (INR Comparable)',
            'Severity': 'Medium',
            'PO No.': row['PO. No.'],
            'Project': row['Project'],
            'Vendor': row['Vendor'],
            'Amount (INR)': row['Difference'],
            'Currency Info': f"Currencies in PRS: {row['currencies']}",
            'Details': f"Difference: ₹{row['Difference']:,.2f} (PO: ₹{row['Final Amt. (INR)']:,.2f}, PRS INR Entry: ₹{row['po_total_inr']:,.2f})"
        })
    
    # Exception 3: Multi-currency POs requiring manual review
    multi_currency_pos = merged[merged['is_multi'] == True]
    for _, row in multi_currency_pos.iterrows():
        severity = 'High' if not row['has_inr'] else 'Low'
        
        if row['has_inr']:
            detail_msg = f"Multi-currency PO with currencies: {row['currencies']}. INR entry exists for comparison."
        else:
            detail_msg = f"Multi-currency PO with currencies: {row['currencies']}. NO INR entry - cannot compare to PO amount (₹{row['Final Amt. (INR)']:,.2f})"
        
        exceptions.append({
            'Exception Type': 'Multi-Currency PO - Review Required',
            'Severity': severity,
            'PO No.': row['PO. No.'],
            'Project': row['Project'],
            'Vendor': row['Vendor'],
            'Amount (INR)': row['Final Amt. (INR)'],
            'Currency Info': row['currencies'],
            'Details': detail_msg
        })
    
    # Exception 4: Pure foreign currency POs - informational
    foreign_only = merged[(merged['has_inr'] == False) & (merged['is_multi'] == False)]
    for _, row in foreign_only.iterrows():
        exceptions.append({
            'Exception Type': 'Foreign Currency PO Only',
            'Severity': 'Info',
            'PO No.': row['PO. No.'],
            'Project': row['Project'],
            'Vendor': row['Vendor'],
            'Amount (INR)': row['Final Amt. (INR)'],
            'Currency Info': row['currency_first'],
            'Details': f"Pure foreign currency: {row['currency_first']} (PRS PO Total: {row['po_total_first']:,.2f}). PO shows ₹{row['Final Amt. (INR)']:,.2f} (converted at PO date). Using PO amount for tracking."
        })
    
    # Exception 5: Overpayment check - ONLY for comparable currencies
    prs_summary_by_po_currency = prs_df.groupby(['PO No.', 'Currency']).agg({
        'Invoice Amount': 'sum',
        'PO Total': 'first'
    }).reset_index()
    
    # Only check overpayment within same currency
    prs_summary_by_po_currency['Overpayment'] = (
        prs_summary_by_po_currency['Invoice Amount'] - 
        prs_summary_by_po_currency['PO Total']
    )
    
    overpayments = prs_summary_by_po_currency[prs_summary_by_po_currency['Overpayment'] > 1000]
    
    for _, row in overpayments.iterrows():
        po_match = po_df[po_df['PO. No.'] == row['PO No.']]
        
        if len(po_match) > 0:
            po_record = po_match.iloc[0]
            project = po_record['Project']
            vendor = po_record['Vendor']
        else:
            prs_match = prs_df[prs_df['PO No.'] == row['PO No.']].iloc[0]
            project = prs_match['Pin/Project'] if 'Pin/Project' in prs_df.columns else 'Unknown'
            vendor = prs_match['Vendor'] if 'Vendor' in prs_df.columns else 'Unknown'
        
        currency = row['Currency']
        currency_symbol = '₹' if currency == 'INR' else currency
        
        exceptions.append({
            'Exception Type': f'Overpayment ({currency})',
            'Severity': 'Critical',
            'PO No.': row['PO No.'],
            'Project': project,
            'Vendor': vendor,
            'Amount (INR)': row['Overpayment'] if currency == 'INR' else None,
            'Currency Info': currency,
            'Details': f"Overpaid by {currency_symbol} {row['Overpayment']:,.2f} (Invoiced: {currency_symbol}{row['Invoice Amount']:,.2f}, PO: {currency_symbol}{row['PO Total']:,.2f})"
        })
    
    # Exception 6: Critical Pending Payments (not approved items)
    today = pd.Timestamp.now()
    critical_pending = prs_df[
        (prs_df['Status'] == 'Pending') & 
        (pd.to_datetime(prs_df['Payment Due Date'], errors='coerce') < today)
    ]
    
    for _, row in critical_pending.iterrows():
        days_overdue = (today - pd.to_datetime(row['Payment Due Date'])).days
        
        if days_overdue > 90:
            severity = 'Critical'
        elif days_overdue > 30:
            severity = 'High'
        else:
            severity = 'Medium'
        
        currency = row.get('Currency', 'Unknown')
        amount = row.get('Invoice Amount', 0)
        currency_symbol = '₹' if currency == 'INR' else currency
        
        exceptions.append({
            'Exception Type': 'Pending Payment Overdue',
            'Severity': severity,
            'PO No.': row['PO No.'],
            'Project': row['Pin/Project'],
            'Vendor': row['Vendor'],
            'Amount (INR)': amount if currency == 'INR' else None,
            'Currency Info': currency,
            'Details': f"Invoice {row['Invoice No.']} pending for {days_overdue} days (Due: {row['Payment Due Date'].strftime('%Y-%m-%d')}) - Amount: {currency_symbol} {amount:,.2f}"
        })
    
    # Exception 7: Rejected Payments requiring review
    rejected = prs_df[prs_df['Status'] == 'Rejected']
    
    for _, row in rejected.iterrows():
        currency = row.get('Currency', 'Unknown')
        amount = row.get('Invoice Amount', 0)
        currency_symbol = '₹' if currency == 'INR' else currency
        
        exceptions.append({
            'Exception Type': 'Rejected Payment',
            'Severity': 'High',
            'PO No.': row['PO No.'],
            'Project': row['Pin/Project'],
            'Vendor': row['Vendor'],
            'Amount (INR)': amount if currency == 'INR' else None,
            'Currency Info': currency,
            'Details': f"Invoice {row['Invoice No.']} rejected - requires investigation. Amount: {currency_symbol} {amount:,.2f}"
        })
    
    # Exception 8: NEW - Inconsistent currency usage within same PO
    po_currency_consistency = prs_df.groupby('PO No.')['Currency'].nunique().reset_index()
    po_currency_consistency.columns = ['PO No.', 'Currency Count']
    inconsistent = po_currency_consistency[po_currency_consistency['Currency Count'] > 2]  # More than 2 currencies
    
    for _, row in inconsistent.iterrows():
        po_no = row['PO No.']
        currencies_used = ', '.join(sorted(prs_df[prs_df['PO No.'] == po_no]['Currency'].unique()))
        
        po_match = po_df[po_df['PO. No.'] == po_no]
        if len(po_match) > 0:
            po_record = po_match.iloc[0]
            project = po_record['Project']
            vendor = po_record['Vendor']
            amount = po_record['Final Amt. (INR)']
        else:
            prs_match = prs_df[prs_df['PO No.'] == po_no].iloc[0]
            project = prs_match.get('Pin/Project', 'Unknown')
            vendor = prs_match.get('Vendor', 'Unknown')
            amount = 0
        
        exceptions.append({
            'Exception Type': 'Multiple Currencies in Single PO',
            'Severity': 'Medium',
            'PO No.': po_no,
            'Project': project,
            'Vendor': vendor,
            'Amount (INR)': amount,
            'Currency Info': currencies_used,
            'Details': f"PO has {row['Currency Count']} different currencies ({currencies_used}). Verify if this is intentional (e.g., partial payments, amendments)."
        })
        
    # ADD THIS AS NEW EXCEPTION TYPE (after Exception 7 in your script)
    # Insert around line 820, before creating the exceptions_df
    
    # Exception 9: PRS Records without PO Number (Data Quality Issue)
    print("  Checking for PRS records without PO numbers...")
    
    blank_po_records = prs_df[
        (prs_df['PO No.'].isna()) | 
        (prs_df['PO No.'] == '') | 
        (prs_df['PO No.'].astype(str).str.strip() == '')
    ]
    
    print(f"  Found {len(blank_po_records)} PRS records with blank PO numbers")
    
    for _, row in blank_po_records.iterrows():
        currency = row.get('Currency', 'Unknown')
        amount = row.get('Invoice Amount', 0)
        status = row.get('Status', 'Unknown')
        currency_symbol = '₹' if currency == 'INR' else currency
        
        # Severity based on status
        if status == 'Approved':
            severity = 'Medium'  # Approved but missing PO link
            detail_msg = f"Payment APPROVED but no PO number linked"
        elif status == 'Pending':
            severity = 'High'  # Pending without PO is concerning
            detail_msg = f"Payment PENDING but no PO number linked"
        elif status == 'Rejected':
            severity = 'Low'  # Rejected, less urgent
            detail_msg = f"Payment REJECTED and no PO number linked"
        else:
            severity = 'Medium'
            detail_msg = f"Payment status '{status}' but no PO number linked"
        
        exceptions.append({
            'Exception Type': 'PRS Record without PO Number',
            'Severity': severity,
            'PO No.': 'BLANK/MISSING',
            'Project': row.get('Pin/Project', 'Unknown'),
            'Vendor': row.get('Vendor', 'Unknown'),
            'Amount (INR)': amount if currency == 'INR' else None,
            'Currency Info': currency,
            'Details': f"{detail_msg}. PRS No: {row.get('PRS No', 'N/A')}, Invoice: {row.get('Invoice No.', 'N/A')}, Amount: {currency_symbol} {amount:,.2f}"
        })
    
    print(f"  ✓ Created {len(blank_po_records)} exceptions for blank PO numbers")
    
    # Create dataframe
    exceptions_df = pd.DataFrame(exceptions)
    
    if len(exceptions_df) == 0:
        exceptions_df = pd.DataFrame({
            'Exception Type': ['No exceptions found'],
            'Severity': ['Info'],
            'PO No.': [''],
            'Project': [''],
            'Vendor': [''],
            'Amount (INR)': [0],
            'Currency Info': [''],
            'Details': ['All records appear normal']
        })
    else:
        # Sort by severity
        severity_order = {'Critical': 0, 'High': 1, 'Medium': 2, 'Low': 3, 'Info': 4}
        exceptions_df['Sort_Order'] = exceptions_df['Severity'].map(severity_order)
        exceptions_df = exceptions_df.sort_values(['Sort_Order', 'Amount (INR)'], ascending=[True, False], na_position='last')
        exceptions_df = exceptions_df.drop('Sort_Order', axis=1)
    
    # Create worksheet
    ws = wb.create_sheet("6_Audit_Exceptions")
    
    for r in dataframe_to_rows(exceptions_df, index=False, header=True):
        ws.append(r)
    
    # Add summary
    data_end_row = len(exceptions_df) + 1
    summary_row = data_end_row + 3
    
    ws[f'A{summary_row}'] = "EXCEPTION SUMMARY BY TYPE & SEVERITY (Currency-Aware)"
    ws[f'A{summary_row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{summary_row}'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws.merge_cells(f'A{summary_row}:E{summary_row}')
    
    summary_row += 2
    ws[f'A{summary_row}'] = "Exception Type"
    ws[f'B{summary_row}'] = "Severity"
    ws[f'C{summary_row}'] = "Count"
    ws[f'D{summary_row}'] = "Total Amount (INR only)"
    ws[f'E{summary_row}'] = "Currencies Affected"
    
    apply_header_style(ws, summary_row)
    
    # Group exceptions with currency info
    if len(exceptions_df) > 1 and 'Exception Type' in exceptions_df.columns:
        exception_summary = exceptions_df.groupby(['Exception Type', 'Severity']).agg({
            'PO No.': 'count',
            'Amount (INR)': lambda x: x.dropna().sum() if len(x.dropna()) > 0 else 0,
            'Currency Info': lambda x: ', '.join(sorted(set(str(i) for i in x.dropna().unique())))
        }).reset_index()
        
        for idx, row in exception_summary.iterrows():
            summary_row += 1
            ws[f'A{summary_row}'] = row['Exception Type']
            ws[f'B{summary_row}'] = row['Severity']
            ws[f'C{summary_row}'] = row['PO No.']
            ws[f'D{summary_row}'] = row['Amount (INR)']
            ws[f'D{summary_row}'].number_format = '#,##0.00'
            ws[f'E{summary_row}'] = row['Currency Info']
    
    # Add currency-specific summary
    currency_summary_row = summary_row + 3
    ws[f'A{currency_summary_row}'] = "CURRENCY-SPECIFIC EXCEPTION BREAKDOWN"
    ws[f'A{currency_summary_row}'].font = Font(bold=True, size=11, color="FFFFFF")
    ws[f'A{currency_summary_row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.merge_cells(f'A{currency_summary_row}:D{currency_summary_row}')
    
    currency_summary_row += 2
    ws[f'A{currency_summary_row}'] = "INR-Comparable Mismatches:"
    ws[f'C{currency_summary_row}'] = f'=COUNTIF(A2:A{data_end_row},"*INR Comparable*")'
    
    currency_summary_row += 1
    ws[f'A{currency_summary_row}'] = "Multi-Currency POs (High Priority):"
    ws[f'C{currency_summary_row}'] = f'=COUNTIFS(A2:A{data_end_row},"Multi-Currency*",B2:B{data_end_row},"High")'
    
    currency_summary_row += 1
    ws[f'A{currency_summary_row}'] = "Multi-Currency POs (Low Priority):"
    ws[f'C{currency_summary_row}'] = f'=COUNTIFS(A2:A{data_end_row},"Multi-Currency*",B2:B{data_end_row},"Low")'
    
    currency_summary_row += 1
    ws[f'A{currency_summary_row}'] = "Foreign Currency Only POs:"
    ws[f'C{currency_summary_row}'] = f'=COUNTIF(A2:A{data_end_row},"Foreign Currency PO Only")'
    
    currency_summary_row += 1
    ws[f'A{currency_summary_row}'] = "Multiple Currencies in Single PO:"
    ws[f'C{currency_summary_row}'] = f'=COUNTIF(A2:A{data_end_row},"Multiple Currencies in Single PO")'
    
    apply_header_style(ws)
    auto_adjust_column_width(ws)
    add_borders(ws)
    
    # Highlight by severity
    critical_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    for row in range(2, data_end_row + 1):
        severity = ws[f'B{row}'].value
        if severity == 'Critical':
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = critical_fill
                ws.cell(row=row, column=col).font = Font(color="FFFFFF", bold=True)
        elif severity == 'High':
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = red_fill
        elif severity == 'Medium':
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = yellow_fill
        elif severity == 'Low':
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = blue_fill
        elif severity == 'Info':
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = green_fill
    
    print(f"  ✓ Found {len(exceptions_df)} exceptions with SMART currency handling")
    
    # Print breakdown
    if len(exceptions_df) > 0:
        critical_count = len(exceptions_df[exceptions_df['Severity'] == 'Critical'])
        high_count = len(exceptions_df[exceptions_df['Severity'] == 'High'])
        medium_count = len(exceptions_df[exceptions_df['Severity'] == 'Medium'])
        
        print(f"  ⚠️ Critical: {critical_count} | High: {high_count} | Medium: {medium_count}")

def analysis_7_contracts_client_summary(contracts_dict, wb):
    """Analysis 7: Live Contracts - Client Summary"""
    print("\nPerforming Analysis 7: Contracts Client Summary...")
    
    all_contracts = []
    
    for sheet_name, df in contracts_dict.items():
        if 'Name of the Clients' in df.columns:
            df_copy = df.copy()
            df_copy['Contract Type'] = sheet_name
            all_contracts.append(df_copy)
    
    if len(all_contracts) == 0:
        print("  No contract data available")
        return
    
    combined = pd.concat(all_contracts, ignore_index=True)
    
    client_summary = combined.groupby('Name of the Clients').agg({
        'Contract Type': lambda x: ', '.join(x.unique()),
        'Purchase Order/Agreement': 'count'
    }).reset_index()
    
    client_summary.columns = ['Client Name (Standardized)', 'Contract Types', 'Number of Contracts']
    client_summary = client_summary.sort_values('Number of Contracts', ascending=False)
    
    ws = wb.create_sheet("7_Contracts_Client_Summary")
    
    for r in dataframe_to_rows(client_summary, index=False, header=True):
        ws.append(r)
    
    data_end_row = len(client_summary) + 1
    summary_row = data_end_row + 2
    
    ws[f'A{summary_row}'] = "TOTAL UNIQUE CLIENTS"
    ws[f'A{summary_row}'].font = Font(bold=True)
    ws[f'C{summary_row}'] = f'=COUNTA(A2:A{data_end_row})'
    
    ws[f'A{summary_row+1}'] = "TOTAL CONTRACTS"
    ws[f'A{summary_row+1}'].font = Font(bold=True)
    ws[f'C{summary_row+1}'] = f'=SUM(C2:C{data_end_row})'
    
    apply_header_style(ws)
    auto_adjust_column_width(ws)
    add_borders(ws)
    
    print(f"  ✓ Analyzed {len(client_summary)} unique clients")
    
def analysis_8_contract_po_utilization(po_df, contracts_dict, wb):
    """
    NEW Analysis 8: Contract vs PO Utilization
    
    Purpose:
    - Link POs to master contracts via Client/Project matching
    - Check if POs exceed contract values
    - Identify POs under closed contracts
    """
    print("\nPerforming Analysis 8: Contract vs PO Utilization...")
    
    # Combine all contract sheets
    all_contracts = []
    for sheet_name, df in contracts_dict.items():
        if 'Name of the Clients' in df.columns:
            df_copy = df.copy()
            df_copy['Contract Type'] = sheet_name
            all_contracts.append(df_copy)
    
    if len(all_contracts) == 0:
        print("  No contract data available")
        return
    
    combined_contracts = pd.concat(all_contracts, ignore_index=True)
    
    # Get relevant columns from contracts
    contract_columns = ['Name of the Clients', 'Purchase Order/Agreement', 
                       'Contract Type', 'Status']
    
    # Try to find contract value column (name may vary)
    value_columns = [col for col in combined_contracts.columns 
                    if any(keyword in str(col).lower() 
                          for keyword in ['value', 'amount', 'contract amt', 'po amt'])]
    
    if value_columns:
        contract_columns.append(value_columns[0])
        combined_contracts[value_columns[0]] = convert_to_numeric(
            combined_contracts[value_columns[0]], value_columns[0]
        )
    
    contracts_summary = combined_contracts[contract_columns].copy()
    contracts_summary.columns = ['Client', 'Contract ID', 'Contract Type', 
                                 'Status', 'Contract Value'] if len(value_columns) > 0 else \
                                ['Client', 'Contract ID', 'Contract Type', 'Status']
    
    # Aggregate POs by client (using standardized vendor names)
    po_by_client = po_df.groupby('Vendor').agg({
        'PO. No.': 'nunique',
        'Final Amt. (INR)': 'sum',
        'Project': lambda x: ', '.join(x.unique()[:3])  # First 3 projects
    }).reset_index()
    
    po_by_client.columns = ['Client', 'Number of POs', 'Total PO Value (INR)', 'Projects']
    
    # Merge contracts with PO data
    analysis = contracts_summary.merge(
        po_by_client,
        on='Client',
        how='left'
    )
    
    analysis['Number of POs'] = analysis['Number of POs'].fillna(0).astype(int)
    analysis['Total PO Value (INR)'] = analysis['Total PO Value (INR)'].fillna(0)
    
    # Calculate utilization if contract value exists
    if 'Contract Value' in analysis.columns:
        analysis['Utilization %'] = np.where(
            analysis['Contract Value'] > 0,
            (analysis['Total PO Value (INR)'] / analysis['Contract Value'] * 100).round(2),
            0
        )
        
        # Flag over-utilization
        analysis['Utilization Flag'] = np.where(
            analysis['Utilization %'] > 100,
            'Over-utilized',
            np.where(
                analysis['Utilization %'] > 90,
                'Near Limit',
                'Within Limit'
            )
        )
    
    # Flag POs under closed contracts
    analysis['Contract Status Flag'] = np.where(
        (analysis['Status'].str.upper() == 'CLOSE') & (analysis['Number of POs'] > 0),
        'POs exist under CLOSED contract',
        'Normal'
    )
    
    # Sort by total PO value
    analysis = analysis.sort_values('Total PO Value (INR)', ascending=False)
    
    # Create worksheet
    ws = wb.create_sheet("8_Contract_PO_Utilization")
    
    for r in dataframe_to_rows(analysis, index=False, header=True):
        ws.append(r)
    
    # Add summary
    data_end_row = len(analysis) + 1
    summary_row = data_end_row + 3
    
    ws[f'A{summary_row}'] = "CONTRACT UTILIZATION SUMMARY"
    ws[f'A{summary_row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{summary_row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells(f'A{summary_row}:E{summary_row}')
    
    summary_row += 2
    ws[f'A{summary_row}'] = "Total Contracts:"
    ws[f'C{summary_row}'] = len(analysis)
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Open Contracts:"
    ws[f'C{summary_row}'] = f'=COUNTIF(D2:D{data_end_row},"Open")'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Closed Contracts:"
    ws[f'C{summary_row}'] = f'=COUNTIF(D2:D{data_end_row},"Close")'
    
    summary_row += 1
    ws[f'A{summary_row}'] = "Contracts with POs:"
    ws[f'C{summary_row}'] = f'=COUNTIF(E2:E{data_end_row},">0")'
    
    if 'Utilization Flag' in analysis.columns:
        summary_row += 2
        ws[f'A{summary_row}'] = "Over-utilized Contracts:"
        ws[f'C{summary_row}'] = f'=COUNTIF(H2:H{data_end_row},"Over-utilized")'
        
        summary_row += 1
        ws[f'A{summary_row}'] = "Near Limit Contracts:"
        ws[f'C{summary_row}'] = f'=COUNTIF(H2:H{data_end_row},"Near Limit")'
    
    summary_row += 2
    ws[f'A{summary_row}'] = "POs under CLOSED Contracts:"
    ws[f'C{summary_row}'] = f'=COUNTIF(I2:I{data_end_row},"POs exist under CLOSED contract")'
    
    apply_header_style(ws)
    auto_adjust_column_width(ws)
    add_borders(ws)
    
    # Highlighting
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    for row in range(2, data_end_row + 1):
        # Highlight over-utilization
        if 'Utilization Flag' in analysis.columns:
            util_flag = ws[f'H{row}'].value
            if util_flag == 'Over-utilized':
                ws[f'H{row}'].fill = red_fill
                ws[f'G{row}'].fill = red_fill
            elif util_flag == 'Near Limit':
                ws[f'H{row}'].fill = yellow_fill
                ws[f'G{row}'].fill = yellow_fill
        
        # Highlight POs under closed contracts
        status_flag = ws[f'I{row}'].value
        if status_flag and 'CLOSED' in str(status_flag):
            ws[f'I{row}'].fill = red_fill
            ws[f'D{row}'].fill = red_fill
    
    print(f"  ✓ Analyzed {len(analysis)} contracts")
    
    # Count issues
    if 'Utilization Flag' in analysis.columns:
        over_utilized = len(analysis[analysis['Utilization Flag'] == 'Over-utilized'])
        print(f"  ⚠️ Over-utilized contracts: {over_utilized}")
    
    closed_with_pos = len(analysis[analysis['Contract Status Flag'] == 'POs exist under CLOSED contract'])
    print(f"  ⚠️ POs under closed contracts: {closed_with_pos}")

def create_executive_summary(po_df, prs_df, contracts_dict, wb):
    """Create Executive Summary Dashboard - SMART CURRENCY-AWARE"""
    print("\nCreating Executive Summary...")
    
    ws = wb.create_sheet("Executive_Summary", 0)
    
    # Title
    ws['A1'] = "INTERNAL AUDIT - EXECUTIVE SUMMARY"
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    ws['A2'] = f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws.merge_cells('A2:D2')
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Key Changes
    row = 4
    ws[f'A{row}'] = "KEY FEATURES - SMART CURRENCY HANDLING"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.merge_cells(f'A{row}:D{row}')
    
    changes = [
        "✓ Smart Currency Detection: Auto-identifies multi-currency POs",
        "✓ INR Priority: Uses INR values for comparison when available",
        "✓ No False Mismatches: Only compares same-currency amounts",
        "✓ Detailed Breakdown: Currency-by-currency transaction details",
        "✓ Manual Review Flags: Clear identification of items needing forex review"
    ]
    
    for change in changes:
        row += 1
        ws[f'A{row}'] = change
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].font = Font(size=10)
    
    # PO Currency Classification
    row += 2
    ws[f'A{row}'] = "PO CURRENCY CLASSIFICATION"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 2
    
    # Calculate currency classification from PRS
    prs_currency_analysis = prs_df.groupby('PO No.').agg({
        'Currency': lambda x: {
            'currencies': list(x.unique()),
            'has_inr': 'INR' in x.values,
            'is_multi': len(x.unique()) > 1,
            'count': len(x.unique())
        }
    })
    
    pure_inr = sum(1 for v in prs_currency_analysis['Currency'] 
                   if v['has_inr'] and not v['is_multi'])
    multi_with_inr = sum(1 for v in prs_currency_analysis['Currency'] 
                         if v['has_inr'] and v['is_multi'])
    foreign_only = sum(1 for v in prs_currency_analysis['Currency'] 
                       if not v['has_inr'] and not v['is_multi'])
    multi_no_inr = sum(1 for v in prs_currency_analysis['Currency'] 
                       if not v['has_inr'] and v['is_multi'])
    
    total_pos_in_prs = len(prs_currency_analysis)
    total_pos_overall = len(po_df['PO. No.'].unique())
    pos_without_prs = total_pos_overall - total_pos_in_prs
    
    ws[f'A{row}'] = "PO Category"
    ws[f'B{row}'] = "Count"
    ws[f'C{row}'] = "% of Total"
    ws[f'D{row}'] = "Status"
    apply_header_style(ws, row)
    
    po_categories = [
        ('Pure INR POs', pure_inr, 'Can compare directly', 'green'),
        ('Multi-Currency with INR', multi_with_inr, 'Can compare using INR entry', 'orange'),
        ('Foreign Currency Only', foreign_only, 'Cannot compare to INR', 'blue'),
        ('Multi-Currency (No INR)', multi_no_inr, 'Manual review required', 'yellow'),
        ('POs without PRS Records', pos_without_prs, 'No payment tracking', 'red')
    ]
    
    for category, count, status, color in po_categories:
        row += 1
        ws[f'A{row}'] = category
        ws[f'B{row}'] = count
        if total_pos_overall > 0:
            ws[f'C{row}'] = f"{(count/total_pos_overall*100):.1f}%"
        else:
            ws[f'C{row}'] = "0.0%"
        ws[f'D{row}'] = status
        
        # Color coding
        color_map = {
            'green': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            'orange': PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid"),
            'blue': PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
            'yellow': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            'red': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        }
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].fill = color_map.get(color, PatternFill())
    
    # Currency breakdown from PRS
    row += 3
    ws[f'A{row}'] = "CURRENCY BREAKDOWN (PRS Transactions)"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 2
    currency_summary = prs_df.groupby('Currency').agg({
        'PRS No': 'count',
        'Invoice Amount': 'sum'
    }).reset_index()
    
    ws[f'A{row}'] = "Currency"
    ws[f'B{row}'] = "Transactions"
    ws[f'C{row}'] = "Total Amount"
    ws[f'D{row}'] = "% of Transactions"
    apply_header_style(ws, row)
    
    total_transactions = currency_summary['PRS No'].sum()
    for _, curr_row in currency_summary.iterrows():
        row += 1
        ws[f'A{row}'] = curr_row['Currency']
        ws[f'B{row}'] = curr_row['PRS No']
        ws[f'C{row}'] = f"{curr_row['Invoice Amount']:,.2f}"
        if total_transactions > 0:
            ws[f'D{row}'] = f"{(curr_row['PRS No']/total_transactions*100):.1f}%"
        else:
            ws[f'D{row}'] = "0.0%"
    
    # Financial Summary (INR only)
    row += 3
    ws[f'A{row}'] = "FINANCIAL SUMMARY (INR Transactions Only)"
    ws[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 2
    po_total = float(po_df['Final Amt. (INR)'].sum())
    
    # Only INR transactions from PRS
    prs_inr = prs_df[prs_df['Currency'] == 'INR']
    prs_invoiced_inr = prs_inr['Invoice Amount'].sum()
    
    # Get PO counts
    total_po_count = len(po_df['PO. No.'].unique())
    
    metrics = [
        ('Total Purchase Orders (All)', total_po_count),
        ('POs with Payment Records', total_pos_in_prs),
        ('POs Comparable (INR)', pure_inr + multi_with_inr),
        ('', ''),
        ('Total PO Amount (INR)', po_total),
        ('Total Invoiced (INR only)', prs_invoiced_inr),
        ('INR Balance Pending', po_total - prs_invoiced_inr),
        ('', ''),
        ('INR Fulfillment %', f"{(prs_invoiced_inr/po_total*100) if po_total > 0 else 0:.2f}%"),
        ('', ''),
        ('Total Payment Records (All currencies)', len(prs_df)),
        ('INR Payment Records', len(prs_inr)),
        ('Foreign Currency Payment Records', len(prs_df) - len(prs_inr)),
    ]
    
    for label, value in metrics:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        if isinstance(value, float) and value > 1000:
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = '₹#,##0.00'
        else:
            ws[f'B{row}'] = value
        row += 1
    
    # Critical Audit Flags
    row += 2
    ws[f'A{row}'] = "CRITICAL AUDIT FLAGS"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 2
    
    # Calculate flags
    pending_overdue = len(prs_df[
        (prs_df['Status'] == 'Pending') & 
        (pd.to_datetime(prs_df['Payment Due Date'], errors='coerce') < pd.Timestamp.now())
    ])
    
    rejected_count = len(prs_df[prs_df['Status'] == 'Rejected'])
    
    # Multi-currency flags
    multi_currency_high = multi_no_inr  # Multi-currency without INR needs review
    multi_currency_info = multi_with_inr  # Multi-currency with INR is informational
    
    flags = [
        ('POs without Payment Records', pos_without_prs, 'High', 'red'),
        ('Pending Payments (Overdue)', pending_overdue, 'Critical', 'darkred'),
        ('Rejected Payments', rejected_count, 'High', 'red'),
        ('Multi-Currency POs (No INR - Manual Review)', multi_currency_high, 'High', 'yellow'),
        ('Multi-Currency POs (Has INR - Track)', multi_currency_info, 'Low', 'blue'),
        ('Foreign Currency Only POs', foreign_only, 'Info', 'lightblue'),
    ]
    
    ws[f'A{row}'] = "Issue"
    ws[f'B{row}'] = "Count"
    ws[f'C{row}'] = "Priority"
    apply_header_style(ws, row)
    
    for flag in flags:
        row += 1
        ws[f'A{row}'] = flag[0]
        ws[f'B{row}'] = flag[1]
        ws[f'C{row}'] = flag[2]
        
        if flag[1] > 0 and flag[3] in ['darkred', 'red', 'yellow']:
            color_map = {
                'darkred': PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
                'red': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
                'yellow': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            }
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].fill = color_map.get(flag[3], PatternFill())
                if flag[3] == 'darkred':
                    ws[f'{col}{row}'].font = Font(color="FFFFFF", bold=True)
    
    # Analysis sheets guide
    row += 3
    ws[f'A{row}'] = "DETAILED ANALYSIS SHEETS"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 2
    sheets_info = [
        ('1', 'PO vs PRS Matching', 'Smart currency-aware financial reconciliation'),
        ('2', 'Vendor Wise Summary', 'Vendor-level analysis (INR only)'),
        ('3', 'Project Wise Summary', 'Project-level spending (INR only)'),
        ('4', 'Payment Status Analysis', 'Payment tracking with original amounts'),
        ('5', 'Tax Analysis', 'Tax verification (INR only)'),
        ('6', 'Audit Exceptions', 'Smart currency-aware exception detection'),
        ('7', 'Contracts Client Summary', 'Client standardization')
    ]
    
    ws[f'A{row}'] = "Sheet"
    ws[f'B{row}'] = "Analysis Name"
    ws[f'C{row}'] = "Description"
    apply_header_style(ws, row)
    
    for sheet_info in sheets_info:
        row += 1
        ws[f'A{row}'] = sheet_info[0]
        ws[f'B{row}'] = sheet_info[1]
        ws[f'C{row}'] = sheet_info[2]
    
    # Important notes
    row += 3
    ws[f'A{row}'] = "IMPORTANT NOTES FOR AUDITORS"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    ws.merge_cells(f'A{row}:D{row}')
    
    notes = [
        "1. Pure INR POs: Can be directly compared and reconciled",
        "2. Multi-Currency with INR: Uses INR entry for comparison, check Currency Breakdown column",
        "3. Foreign Currency Only: Cannot compare to INR PO amount - use PO amount directly",
        "4. Multi-Currency (No INR): Requires manual review with actual forex rates",
        "5. Sheet 1 Column 'PO Total (For Comparison)': Shows which PRS value is used for matching",
        "6. Sheet 1 'Currency Matching Status': Quick filter for reviewable vs manual review items",
        "7. Sheet 6: Exception detection now currency-aware - no false mismatches"
    ]
    
    for note in notes:
        row += 1
        ws[f'A{row}'] = note
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
    
    auto_adjust_column_width(ws)
    add_borders(ws, start_row=1, end_row=row)
    
    print("  ✓ Executive summary created with SMART currency handling")
    print(f"  ✓ Pure INR POs: {pure_inr} | Multi-Currency with INR: {multi_with_inr}")
    print(f"  ⚠️ Foreign Only: {foreign_only} | Multi-Currency (No INR): {multi_no_inr}")

# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main execution function"""
    print("="*80)
    print("INTERNAL AUDIT ANALYSIS - UPDATED VERSION")
    print("NO CURRENCY CONVERSION - Using PO amounts directly for foreign currency")
    print("="*80)
    
    try:
        # Load data
        po_df = load_po_data()
        prs_df = load_prs_data()
        contracts_dict = load_contracts_data()
        
        # Validation
        print("\nData Validation:")
        print(f"  PO Data: {len(po_df)} records, {po_df['PO. No.'].nunique()} unique POs")
        print(f"  PRS Data: {len(prs_df)} records, {prs_df['PO No.'].nunique()} unique POs")
        
        # Currency breakdown
        currency_counts = prs_df['Currency'].value_counts()
        print("\nCurrency Distribution in PRS:")
        for currency, count in currency_counts.items():
            print(f"  {currency}: {count} transactions")
        
        # Create workbook
        print(f"\nCreating output workbook: {OUTPUT_FILE}")
        wb = Workbook()
        wb.remove(wb.active)
        
        # Perform analyses
        analysis_1_po_vs_prs_matching(po_df, prs_df, wb)
        analysis_2_vendor_wise_summary(po_df, prs_df, wb)
        analysis_3_project_wise_summary(po_df, prs_df, wb)
        analysis_4_payment_status_tracking(prs_df, wb)
        analysis_5_tax_analysis(po_df, prs_df, wb)
        analysis_6_audit_exceptions(po_df, prs_df, wb)
        analysis_7_contracts_client_summary(contracts_dict, wb)
        analysis_8_contract_po_utilization(po_df, contracts_dict, wb)  # NEW
        
        # Create executive summary
        create_executive_summary(po_df, prs_df, contracts_dict, wb)
        
        # Save
        wb.save(OUTPUT_FILE)
        
        print("\n" + "="*80)
        print("✓ AUDIT ANALYSIS COMPLETED SUCCESSFULLY!")
        print("="*80)
        print(f"\n📊 Output file: {OUTPUT_FILE}")
        print("\n🔧 KEY CHANGES:")
        print("  ✓ NO currency conversion applied")
        print("  ✓ Foreign currency POs use PO amounts directly")
        print("  ✓ Column renamed: 'PO Total AsPer PRS_Record'")
        print("  ✓ INR-only financial calculations")
        print("  ✓ Foreign currency transactions clearly identified")
        print("\n📋 Review Priority:")
        print("  1. Sheet 6: Audit Exceptions (Critical & High severity)")
        print("  2. Sheet 4: Payment Status Analysis (Pending overdue items)")
        print("  3. Sheet 1: PO vs PRS Matching (check foreign currency POs)")
        print("\n💡 Note: Foreign currency POs are analyzed separately - no INR conversion")
        print("="*80)
        
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()